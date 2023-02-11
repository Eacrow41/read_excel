from datetime import datetime
from tkinter import ttk
from tkinter import *
import tkinter as tk 
import os
import time
from xml.dom.minidom import Element
import pandas as pd
from openpyxl import load_workbook
import pandas.io.formats.excel
import xlrd
import numpy as np
import json
from openpyxl. styles import Alignment
from pandas import array

class Product:
    # DEFINIMOS EL CONSTRUCTOR Y TOMAMOS EL PARAMETRO QUE NOS ESTAN PASANDO AL INICIAR LA VENTANA
    def __init__(self, window):

        # CREAR PROPIEDAD U OBJETO PARA ALMACENAR LA VENTANA QUE ESTA RECIBIENDO EL CONSTRUCTOR
        window['bg'] = '#B5AEB4'
        ancho_ventana = 990
        alto_ventana = 650

        x_ventana = window.winfo_screenwidth() // 2 - ancho_ventana // 2
        y_ventana = window.winfo_screenheight() // 2 - alto_ventana // 2

        posicion = str(ancho_ventana) + "x" + str(alto_ventana) + "+" + str(x_ventana) + "+" + str(y_ventana)
        window.geometry(posicion)
        window.resizable(0,0)
        self.win = window
        self.win.title('Aplicacion de Productos')

        self.interfaz_inicial1()
        self.interfaz_inicial2()
       
        # LABEL PARA MENSAJES DE SALIDA
        self.message = Label(text = '')
        self.message['bg'] = '#B5AEB4'
        self.message.grid(row = 3, column = 1, columnspan = 1, padx = 20, sticky = W + E)
        
        # TABLA
        frame2 = LabelFrame(self.win, text='Lista de directorios:')
        frame2.grid(row = 1, column = 2, columnspan = 2, sticky= W + E)
        # Input Nombre Buscar
        Label(frame2, text = 'Nombre :').grid(row = 1, column = 0, padx = 20, pady = 5, sticky = W)
        self.buscar = UpperEntry(frame2, width=35)
        self.buscar.grid(row = 1 , column = 1)
        self.buscar.focus()
       

        #Button Agregar Comentario
        ttk.Button(frame2, text = 'Buscar', command=self.find_product).grid(row= 1, column= 2, columnspan=1, sticky= W + E)
        ttk.Button(frame2, text = 'Limpiar', command=self.del_find).grid(row= 1, column= 3, columnspan=1, sticky= W + E)

        self.tree = ttk.Treeview(frame2, height = 10, columns=("#1"))
        self.tree.grid(row = 2, column = 1, columnspan = 3)
        self.tree.heading("#0", text="Fecha", anchor = CENTER)         
        self.tree.heading("#1", text="Nombre", anchor = CENTER)

        # BOTONES
        ttk.Button(text='+', command=self.edit_product).grid(row = 2, column = 2, columnspan=1)
        ttk.Button(text='-', command=self.del_product).grid(row = 2, column = 3, columnspan=1)


        self.get_products()

    # CREACION INTERFAZ GRAFICA
    def interfaz_inicial1(self):
        global frame
        # FRAME CONTENEDOR
        frame = LabelFrame(self.win, text=' Registrar Nuevo Comentario: ')
        frame.grid(row = 1, column = 1, columnspan = 1, pady = 20, padx = 20, ipadx = 50, ipady = 20)

        # INPUT Tipo Actividad
        Label(frame, text=' Tipo Actividad: ').grid(row = 1, column = 0, padx = 5, pady = 5, sticky = W)
        self.tactividad = UpperEntry(frame, width=35)
        self.tactividad.grid(row = 1, column = 1, sticky = W)

        # INPUT Obra
        Label(frame, text=' Obra: ').grid(row = 2, column = 0, padx = 5, pady = 5, sticky = W)
        self.obra = UpperEntry(frame, width=35)
        self.obra.grid(row = 2, column = 1, sticky = W)

        # INPUT usuario
        Label(frame, text=' Usuario: ').grid(row = 3, column = 0, padx = 5, pady = 5, sticky = W)
        self.usuario = UpperEntry(frame, width=35)
        self.usuario.grid(row = 3, column = 1, sticky = W)

        # INPUT Actividad
        Label(frame, text=' Actividad: ').grid(row = 4, column = 0, padx = 5, pady = 5, sticky = W)
        self.actividad = UpperEntry(frame, width=35)
        self.actividad.grid(row = 4, column = 1, sticky = W)

        # INPUT Fecha
        Label(frame, text=' Fecha: ').grid(row = 5, column = 0, padx = 5, pady = 5, sticky = W)
        self.fecha = UpperEntry(frame, width=35)
        self.fecha.grid(row = 5, column = 1, sticky = W)

        # INPUT Encargado
        Label(frame, text=' Encargado: ').grid(row = 6, column = 0, padx = 5, pady = 5, sticky = W)
        self.encargado = UpperEntry(frame, width=35)
        self.encargado.grid(row = 6, column = 1, sticky = W)

        # INPUT Observaciones
        Label(frame, text=' Observaciones: ').grid(row = 7, column = 0, padx = 5, pady = 5, sticky = W)
        self.observacion = UpperEntry(frame, width=35)
        self.observacion.grid(row = 7, column = 1, sticky = W)

     # CREACION INTERFAZ GRAFICA
    def interfaz_inicial2(self):
        global frame
        # FRAME CONTENEDOR
        frame = LabelFrame(self.win, text='Modifica directorios: ')
        frame.grid(row = 3, column = 2, columnspan = 2, pady = 1, padx = 20, ipadx = 10, ipady=10)
    

        self.treeTwo = ttk.Treeview(frame, height = 10, columns=("#1"))
        self.treeTwo.grid(row = 4, column = 1, columnspan = 3, padx = 20)
        self.treeTwo.heading("#0", text="Fecha", anchor = CENTER)         
        self.treeTwo.heading("#1", text="Nombre", anchor = CENTER)

         # BOTON AGREGAR PRODUCTO
        ttk.Button(self.win, text='Agregar Comentario', command=self.add_product).grid(row = 2, column = 1, columnspan = 1, ipadx=20, pady = 10)

    # OBTENER DATOS DE LA TABLA product
    def get_products(self):

        #limpiando treeview
        records = self.tree.get_children()
        for element in records:
            self.tree.delete(element)

        with open('config.txt') as f:
         lines = json.load(f)

        dir = lines['ruta']
        with os.scandir(dir) as ficheros:
            for fichero in ficheros:
                ti_m = os.path.getmtime(fichero) 
                m_ti = time.ctime(ti_m) 
                t_obj = time.strptime(m_ti) 
                T_stamp = time.strftime("%Y-%m-%d-%H:%M:%S", t_obj) 
                self.tree.insert("" , END,  text= T_stamp, values=(fichero.name))

    # VALIDAR CAMPOS
    def validation(self):
            return len(self.tactividad.get()) != 0  and len(self.obra.get()) != 0 and len(self.usuario.get()) != 0 and len(self.actividad.get()) != 0  and len(self.fecha.get()) != 0 and len(self.encargado.get()) != 0 and len(self.observacion.get()) != 0 
          
    # AGREGAR PRODUCTOS
    def add_product(self):
        if self.validation():
            
            try:
                archivos = []
                with open('config.txt') as f:
                    lines = json.load(f)

                direc = lines['ruta']
                archivo_id = self.treeTwo.get_children()
                if len(archivo_id) != 0:
                    for elemnt in archivo_id:
                        text_1 = self.treeTwo.item(elemnt, option="values")
                        archivos.append(text_1[0])
                    
                    
                    for archivo_name in archivos:
                            indice_c = archivo_name.index('-')
                            indice_v = int(indice_c) + 4
                            if indice_v == ('-', '+'):
                                indice_c = int(indice_c) + 1
                                nombre_archivo_corto = archivo_name[indice_c:10]
                            else :
                                indice_c = int(indice_c) + 1
                                nombre_archivo_corto = archivo_name[indice_c:11]

                            
                            nombre_archivo = "CONTROL DE MANTENIMIENTO "+nombre_archivo_corto+".xlsx"
                            archivo = direc+archivo_name+"/"+ nombre_archivo
                            
                            wb = xlrd.open_workbook(archivo)
                            hoja = wb.sheet_by_name(nombre_archivo_corto) 
                            for i in range(0,hoja.nrows):
                                celda = i
                            celda = celda + 2
                          # self.encargado.get(), self.observacion.get()                       
                            # Configuramos Pandas y cargamos el archivo correspondiente                    
                            wb = load_workbook(archivo)

                            # Seleccionamos el archivo
                            sheet = wb.active

                            # Ingresamos el valor 56 en la celda 'A1'
                            celda1 = 'A' + str(celda)
                            sheet[celda1] = self.tactividad.get()
                            celda2 = 'B' + str(celda)
                            sheet[celda2] = self.obra.get()
                            celda3 = 'C' + str(celda)
                            sheet[celda3] = self.usuario.get()

                            celda4 = 'D' + str(celda) + ':' + 'F' + str(celda)   
                            sheet.merge_cells(celda4)
                            celda8 = 'D' + str(celda)
                            sheet[celda8] =  self.actividad.get()
                            cell = sheet.cell(row=1, column=1)
                            cell.value = self.actividad.get()
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                            celda5 = 'G' + str(celda)
                            sheet[celda5] = self.fecha.get()

                            celda6 = 'H' + str(celda) + ':' + 'I' + str(celda)   
                            sheet.merge_cells(celda6) 
                            celda9 = 'H' + str(celda)
                            sheet[celda9] = self.encargado.get()                           
                            cell = sheet.cell(row=1, column=1)
                            cell.value = self.encargado.get()
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                            celda7 = 'J' + str(celda) + ':' + 'K' + str(celda)   
                            sheet.merge_cells(celda7)
                            celda10 = 'J' + str(celda)
                            sheet[celda10] = self.observacion.get()
                            cell = sheet.cell(row=1, column=1)
                            cell.value = self.observacion.get()
                            cell.alignment = Alignment(horizontal='center', vertical='center')


                            # Guardamos el archivo con los cambios
                            wb.save(archivo)
                            f = open('log.txt','a')
                            f.write('\n' + 'Directorio Modificado: ' + nombre_archivo + ' '+ str(datetime.now()))
                            f.close()

                    
                    self.message['text'] = 'Comentario agregado!!'
                    self.message['font'] = ('Consolas',11)
                    self.message['bg'] ='#d4edda'
                    self.message['fg'] ='#116158'

                    # LIMPIAR CAMPOS
                    self.get_products()
                    self.limpiar_tree()
                    self.buscar.delete(0, END)
                    self.tactividad.delete(0, END)
                    self.obra.delete(0, END)
                    self.usuario.delete(0, END)
                    self.actividad.delete(0, END)
                    self.fecha.delete(0, END)
                    self.encargado.delete(0, END)
                    self.observacion.delete(0, END)
                else:

                    self.message['text'] = 'Debe seleccionar por lo menos un archivo para agregar'
                    self.message['font'] = ('Consolas',11)
                    self.message['bg'] ='#f7d7da'
                    self.message['fg'] ='#89312f'  

            except IndexError as e:
                self.message['text'] = 'El proceso no fue exitoso por favor revisar el archivo Log'
                self.message['font'] = ('Consolas',11)
                self.message['bg'] ='#f7d7da'
                self.message['fg'] ='#89312f'
                f = open('log.txt','a')
                f.write('\n' + 'Directorio Modificado: ' + nombre_archivo + ' '+ str(datetime.now())+ ' '+IndexError)
                f.close()

        else:
            self.message['text'] = 'Todos los campos son requeridos...'
            self.message['font'] = ('Consolas',11)
            self.message['bg'] ='#f7d7da'
            self.message['fg'] ='#89312f'

    # ELIMINAR PRODUCTOS
    def del_product(self):
        self.message['text'] = ''
        nuevo = Entry()
        try:
            self.treeTwo.item(self.treeTwo.selection())['text'][0]
            self.message['bg'] ='#B5AEB4'
        except IndexError as e:
            self.message['text'] = 'Por favor selecciona un registro'
            self.message['font'] = ('Consolas',11)
            self.message['bg'] ='#f7d7da'
            self.message['fg'] ='#89312f'

            return
        selected_item = self.treeTwo.selection()[0]
        self.treeTwo.delete(selected_item)
    
    # EDITAR PRODUCTOS
    def edit_product(self):
        self.message['text'] = ''
        try:
            self.tree.item(self.tree.selection())['text'][0]
            self.message['bg'] ='#B5AEB4'
        except IndexError as e:
            self.message['text'] = 'Por favor selecciona un registro'
            self.message['font'] = ('Consolas',11)
            self.message['bg'] ='#f7d7da'
            self.message['fg'] ='#89312f'
            return
        
        name1 = self.tree.item(self.tree.selection())['text']
        name2 = self.tree.item(self.tree.selection())['values']
        inserta = 'true'
        archivo_id = self.treeTwo.get_children()
        for elemnt in archivo_id:
            text_1 = self.treeTwo.item(elemnt, option="values")
            if text_1[0] == name2[0]: 
                inserta = 'false'
        #archivos_edit.append(name2)
        if inserta == 'true':
            self.treeTwo.insert("" , END,  text= name1, values=(name2))
            
        self.del_find()

    def limpiar_tree(self):
        records = self.treeTwo.get_children()
        for elemnt in records:
            self.treeTwo.delete(elemnt)
    
    # BUSCAR PRODUCTOS
    def find_product(self):
        if  len(self.buscar.get()) != 0 :
            with open('config.txt') as f:
                lines = json.load(f)
            self.message['text'] = ''
            dir = lines['ruta']
            directorio = dir+self.buscar.get()
            if os.path.exists(directorio):
                self.message['bg'] ='#B5AEB4'
                archivo_id = self.tree.get_children()
                if len(archivo_id) > 1:
                    for elemnt in archivo_id:
                        name1 = self.tree.item(elemnt, option="text")
                        name2 = self.tree.item(elemnt, option="values")
                        if self.buscar.get() == name2[0]:
                            fecha = name1
                            nombre_dir = name2
                            
                    #limpiando treeview
                    records = self.tree.get_children()
                    for element in records:
                        self.tree.delete(element) 

                    self.tree.insert("" , END,  text= fecha, values=(nombre_dir)) 

                else:
                    self.get_products()
                    archivo_id = self.tree.get_children()
                    for elemnt in archivo_id:
                        name1 = self.tree.item(elemnt, option="text")
                        name2 = self.tree.item(elemnt, option="values")
                        if self.buscar.get() == name2[0]:
                            fecha = name1
                            nombre_dir = name2
                     #limpiando treeview
                    records = self.tree.get_children()
                    for element in records:
                        self.tree.delete(element) 

                    self.tree.insert("" , END,  text= fecha, values=(nombre_dir)) 
            
            else:
                self.message['text'] = 'No se encontro el directorio'
                self.message['font'] = ('Consolas',11)
                self.message['bg'] ='#f7d7da'
                self.message['fg'] ='#89312f'
                return
        else:
                self.message['text'] = 'El campo Nombre es obligatorio'
                self.message['font'] = ('Consolas',11)
                self.message['bg'] ='#f7d7da'
                self.message['fg'] ='#89312f'
                self.buscar.focus()
    
    # ELIMINAR BUSCADOR
    def del_find(self):
        #limpiando treeview
        records = self.tree.get_children()
        for element in records:
             self.tree.delete(element) 

        self.message['text'] = ''
        self.message['bg'] ='#B5AEB4'

        self.buscar.delete(0, END)
        self.get_products()
        

class UpperEntry(Entry):
    def __init__(self, parent, *args, **kwargs):
        self._var = kwargs.get("textvariable") or StringVar(parent)
        super().__init__(parent, *args, **kwargs)
        self.configure(textvariable=self._var)
        self._to_upper()

    def config(self, cnf=None, **kwargs):
        self.configue(cnf, **kwargs)

    def configure(self, cnf=None, **kwargs):
        var = kwargs.get("textvariable")
        if var is not None:
            var.trace_add('write', self._to_upper)
            self._var = var
        super().config(cnf, **kwargs)

    def __setitem__(self, key, item):
        if key == "textvariable":
            item.trace_add('write', self._to_upper)
            self._var = item
        super.__setitem__(key, item)

    def _to_upper(self, *args):
        self._var.set(self._var.get().upper())


if __name__ == "__main__":
    window = Tk()
    aplication = Product(window)
    window.mainloop()